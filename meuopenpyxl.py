from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

data = {
	"Joe": {
		"math": 65,
		"science": 78,
		"english": 98,
		"gym": 89
	},
	"Bill": {
		"math": 55,
		"science": 72,
		"english": 87,
		"gym": 95
	},
	"Tim": {
		"math": 100,
		"science": 45,
		"english": 75,
		"gym": 92
	},
	"Sally": {
		"math": 30,
		"science": 25,
		"english": 45,
		"gym": 100
	},
	"Jane": {
		"math": 100,
		"science": 100,
		"english": 100,
		"gym": 60
	}
}

wb = Workbook()
ws = wb.active
ws.title = "Grades"

headings = ['Name'] + list(data['Joe'].keys())
ws.append(headings)

for person in data:
	grades = list(data[person].values())
	ws.append([person] + grades)

for col in range(2, len(data['Joe']) + 2):
	char = get_column_letter(col)
	ws[char + "7"] = f"=SUM({char + '2'}:{char + '6'})/{len(data)}"

for col in range(1, 6):
	ws[get_column_letter(col) + '1'].font = Font(bold=True, color="0099CCFF")

wb.save("NewGrades.xlsx")

# wb = load_workbook('teste.xlsx') # Abrir arquivo Excel
# ws = wb.active # Abrir aba atual
# ws.title = 'Dados' # Alterar o nome da aba
# ws.merge_cells('A1:D1') # Mesclar células
# ws.unmerge_cells('A1:D1') # Desmesclar células
# ws.insert_rows(7) # Inserir linha
# ws.delete_rows(7) # Deleta linha
# ws.insert_cols(2) # Inserir coluna
# ws.delete_cols(2) # Deleta coluna
# ws.move_range('C1:D11', rows=2, cols=2) # Movimentar um grupo de células
#wb.create_sheet('teste2') # Criar uma nova aba
#ws = wb['teste2'] # Abrir a aba específica
#ws['a1'].value = 'teste...' # Atribuir valor
#ws.append(['Isaias','é','muito','legal!'])
#print(wb.sheetnames) # Nome das abas do arquivo
# for row in range(1, 11):
#     for col in range(1, 5):
#         #char = chr(65 + col) forma manual o range muda para (0, 4)
#         char = get_column_letter(col)
#         print(ws[char + str(row)], '-', ws[char + str(row)].value)
#         ws[char + str(row)] = char + str(row)
#s = input('Name?')
#wb.save(s+'.xlsx') # Salvar um arquivo